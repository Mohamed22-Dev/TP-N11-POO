import tkinter as tk
from tkinter import messagebox
import os
try:
    from openpyxl import Workbook, load_workbook
except:
    messagebox.showerror("Erreur", "Installe openpyxl avec : pip install openpyxl")
    exit()

# ===== Fenêtre =====
root = tk.Tk()
root.title("Form d'inscription")
root.geometry("500x400")

labels = ["Nom", "Prénom", "Date de naissance", "Téléphone", "Email", "Adresse", "Niveau"]
entries = []

# ===== Form =====
for i, text in enumerate(labels):
    tk.Label(root, text=text).grid(row=i, column=0, padx=10, pady=5)

    entry = tk.Entry(root, width=40)
    entry.grid(row=i, column=1, padx=10, pady=5)
    entries.append(entry)

# ===== Fonction =====
def save_data():
    data = [e.get() for e in entries]

    if "" in data:
        messagebox.showwarning("Erreur", "Remplis tous les champs")
        return

    file = "stagiaires.xlsx"

    if os.path.exists(file):
        wb = load_workbook(file)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(labels)

    ws.append(data)
    wb.save(file)

    messagebox.showinfo("Succès", "Enregistré !")

    for e in entries:
        e.delete(0, tk.END)

# ===== Bouton =====
tk.Button(root, text="Enregistrer", command=save_data).grid(row=len(labels), column=1, pady=20)

root.mainloop()